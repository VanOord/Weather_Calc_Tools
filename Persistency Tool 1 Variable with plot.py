## Finished Aug. 17th, 2023
## Author: q4r, Property of Van Oord Offshore Wind
import pandas as pd
import numpy as np
import calendar
import PySimpleGUI as sg
import plotly.graph_objects as go
from dateutil.parser import parse
import openpyxl
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side, PatternFill
from scipy.stats import norm
# Defining functions

def find_data_start(file_path): #Function skips any rows minus 1 until the datetime data starts for atleast 20 rows
    with open(file_path) as file:
        lines = file.readlines()

    for i, line in enumerate(lines):
        row_data = line.strip().split(",")
        if len(row_data) >= 2:
            try:
                parse(row_data[0])
                if all(len(line.strip().split(",")) >= 2 for line in lines[i+1:i+21]):
                    return i - 1  # Adjusted to return the index of the header line
            except ValueError:
                continue

    return None
def calculate_persistency(values, window_size, limit1, values_op): #Calculates persistency for the overlapping case
    num_windows = len(values) - window_size + 1
    window_values = values[np.arange(window_size)[:, None] + np.arange(num_windows)]
    threshold_met = (window_values <= limit1) if values_op == '-RADIO BELOW-' else (window_values >= limit1)
    count = np.sum(np.all(threshold_met, axis=0)) # This value gives the actual occurence number of overlapping weather windows (Not used as an output currently)
    persistency = (count / len(values)) * 100 # Gives persistency percentage
    return persistency, count


def calculate_persistency_non_overlapping(values, window_size, limit1, values_op): #Calculates persistency for the non-overlapping case
    if len(values) == 0:
        return 0, 0  # Avoid division by zero and return zeros if values list is empty

    values = np.array(values)  # Convert to numpy array for element-wise operations
    count = 0
    i = 0
    
    while i <= len(values) - window_size:
        window = values[i:i+window_size]
        
        if values_op == '-RADIO BELOW-':
            condition_met = np.all(window <= limit1)
        else:
            condition_met = np.all(window >= limit1)
        
        if condition_met:
            count += 1
            i += window_size  # move to the end of this window
        else:
            i += 1  # move one step forward
    
    #max_possible_windows = len(values)//window_size
    #print(max_possible_windows) #Troubleshooting
    #persistency = (count / max_possible_windows) * 100 if max_possible_windows != 0 else 0 #userful if you want to calculate duration time
    persistency = count/len(values) * 100
    return persistency, count


def main_calculations(file_path, time, investigation1):
    
    output_messages = [] #create GUI output message dictionary

    # Determine the start of the data based on if there is a datetime format atleast 20 rows long in the first column
    data_start = find_data_start(file_path)
    if data_start is None:
        output_messages.append("Data start not found.")
        return

    # Load chunk of data from the CSV file to get the header names of the columns
    chunk_df = pd.read_csv(file_path, nrows=100, skiprows=data_start)
    time = chunk_df.columns[0] #will be the index column, CODE ONLY WORKS IF DATETIME COLUMN IS FIRST
    header_names1 = list(chunk_df.columns[1:]) #Skips first column header (time)

    # Load next chunk of data to get the sampling rate in minutes
    sample_df = pd.read_csv(file_path, nrows=1000, skiprows=data_start, usecols=[time], parse_dates=[time], dayfirst=True)
    sample_df.set_index(time, inplace=True)
    time_diffs = sample_df.index.to_series().diff().dropna()  # Calculate differences between consecutive datetime entries
    inferred_sample_rate = time_diffs.mode().iloc[0] #Inferred sample rate in a 'Timedelta' spread
    inferred_minutes = inferred_sample_rate.seconds//60 #gets the sample rate in minute from the number of seconds

    #If statement to get correct labels and sample rate to output in GUI
    if inferred_minutes < 60:
        sample_rate_hm = int(inferred_minutes)
        hours_or_minute = 'minutes'
        window_input_label = 'minutes'
        sample_rate_type = 'minutes'
    elif inferred_minutes == 60:
        sample_rate_hm = int(inferred_minutes/60)
        hours_or_minute = 'hours'
        sample_rate_type = 'hour'
        window_input_label = 'hour(s)'
    else:
        sample_rate_hm = inferred_minutes/60
        hours_or_minute = 'hours'
        sample_rate_type = 'hours'
        window_input_label = 'hour(s)'
    # Define the layout for the GUI window
    layout = [
        #[sg.Text(f"Sample rate: {inferred_minutes} minutes")],
        [sg.Text("Choose column to investigate: ")],
        [sg.Text("Column 1 (required): "), sg.Combo(header_names1, default_value=investigation1, key='investigation1'), sg.Text("Limit: "), sg.Input(key='-LIMIT1-', size=(15,1))],
        #[sg.Text('')], #can add a spaced line here if wanted
        #Other definitions
        [sg.Radio('Calculate less than and equal to limit', 'RADIO', key='-RADIO BELOW-', default=True),
         sg.Radio('Calculate greater than and equal to limit', 'RADIO', key='-RADIO ABOVE-')],
        [sg.Text(f"Enter weather window duration in {hours_or_minute} which are integer divisible by the sample rate of {sample_rate_hm} {sample_rate_type}:")], #(Enter '1' for default timestep of {inferred_minutes} minutes)
        [sg.Input(key='-WINDOW SIZE-', size=(5,1)), sg.Text(f"{window_input_label}     (Enter {sample_rate_hm} to equal timestep)")],
        #[sg.Text('')],
        [sg.Text('Weather Window Type:')],
        [sg.Radio('Overlapping', 'RADIO3', key='-RADIO overlapping-', default=True), sg.Radio('Non-overlapping', 'RADIO3', key='-RADIO NON overlapping-')],
        #[sg.Text('')],
        [sg.Text('If desired, enter P-value integer besides given P20, P50, and P80: P'), sg.Input(key='-PVALUE-', size=(15,1))],
        [sg.Text('Do you want to investigate persistency based on certain months?')],
        [sg.Radio('No', 'RADIO1', key='-RADIO NO-', default=True), sg.Radio('Yes', 'RADIO1', key='-RADIO YES-')],
        #[sg.Text('')],
        [sg.Text('              Start month:'), sg.Combo(list(calendar.month_name[1:]), key='-START MONTH-', enable_events=True, disabled=False)],
        [sg.Text('              End month:'), sg.Combo(list(calendar.month_name[1:]), key='-END MONTH-', enable_events=True, disabled=False)],
        #[sg.Text('')],
        [sg.Button("Plot"), sg.Button("Exit")]
    ]

    # Create the GUI window
    window = sg.Window("Persistency Tool", layout)

    # Event loop to process GUI events
    while True:
        event, values = window.read()

        if event == "Plot":
            investigation1 = values['investigation1']
            limit1 = float(values['-LIMIT1-'])
            values_op = '-RADIO BELOW-' if values['-RADIO BELOW-'] else '-RADIO ABOVE-'
            window_size = int(values['-WINDOW SIZE-'])
            pvalue_input = int(values['-PVALUE-']) if values['-PVALUE-'] else None
            overlapping_WW = values['-RADIO overlapping-']
            investigate_range = values['-RADIO YES-']
            start_month = values['-START MONTH-']
            end_month = values['-END MONTH-']
            window.close()
            break

        if event in (sg.WINDOW_CLOSED, "Exit"):
            break

    # Close the GUI window
    window.close()
    
    # Create a list of columns to load based on user's input
    columns_to_load = [time, investigation1]
    
    #Load all the chosen data
    df = pd.read_csv(file_path, skiprows=data_start, usecols=columns_to_load, parse_dates=[time], dayfirst=True)
    #Set Index columnn as the first column
    df.set_index(time, inplace=True)
    #This loop specifies if the output line displays minutes or hour(s) and makes sure the window_size is correct integer for the persistency calculations
    if inferred_minutes < 60:
        window_time = window_size
        window_size = window_size // inferred_minutes
    else:
        window_time = window_size
        window_size = window_size

    # Resampling the data by month for the mean count in the non-overlapping for all data case
    monthly_counts_nonfiltered = df.resample('M').size()

     # Wiggle room in hours (still will consider a month full even if it is missing upto 13 hours)
    wiggle_room = 13

    # Determine which months have a count close to the full number of hours for that month
    full_months_approx = monthly_counts_nonfiltered[monthly_counts_nonfiltered.index.map(lambda dt: calendar.monthrange(dt.year, dt.month)[1]*24 - wiggle_room) <= monthly_counts_nonfiltered]

    # Get the number of full months considering the wiggle room
    num_full_months_approx = len(full_months_approx)
    
     # Create the figure that will display the chosen variable over time
    fig = go.Figure()


    #calculation of persistency percentage and count
    if investigate_range: #this case is for if certain months want to be investigated
        month_dict = {v.lower(): k for k, v in enumerate(calendar.month_name) if k != 0}
        start_month_num = month_dict[start_month.lower()]
        end_month_num = month_dict[end_month.lower()]
        selected_months = range(start_month_num, end_month_num + 1)
        df_filtered = df[df.index.month.isin(selected_months)]
        values1 = df_filtered[investigation1].values
        years = df.index.year.unique()
        complete_ranges_count = 0

        for year in years:
            # Filter dataframe for the specific year
            df_year = df[df.index.year == year]
            # Check if all months in the selected range exist for that year
            if set(selected_months).issubset(set(df_year.index.month)):
                complete_ranges_count += 1

        if overlapping_WW:
            persistency, count = calculate_persistency(values1, window_size, limit1, values_op)
            additional_info_parts = [f"Persistency: {persistency:.2f}% for months {start_month} through {end_month} with an overlapping weather window of {window_time} {window_input_label} and a limit of {investigation1} {('<=' if values_op == '-RADIO BELOW-' else '>=')} {limit1}"]
        else:
            persistency, count = calculate_persistency_non_overlapping(values1, window_size, limit1, values_op)
            mean_count = count/complete_ranges_count #Mean number of WW occurrences per year
            additional_info_parts = [f"Persistency: {persistency:.2f}% for months {start_month} through {end_month} with a non-overlapping weather window of {window_time} {window_input_label} and a limit of {investigation1} {('<=' if values_op == '-RADIO BELOW-' else '>=')} {limit1}"]

        if overlapping_WW == False:
            # Append the additional provided content
            additional_info_parts.append("\n" + "\n" + f"Total number of weather window occurences in chosen range: {count:.2f}" + "\n"+ f"Mean number of weather window occurrences per chosen range: {mean_count:.2f} ")

        # Join the parts to create the full additional info string
        additional_info = "; ".join(additional_info_parts)

        output_messages.append(additional_info)
        output_messages.append("")


    else: #this case is for when all the data wants to be looked at
        values1 = df[investigation1].values

        if overlapping_WW:
            persistency, count = calculate_persistency(values1, window_size, limit1, values_op)
            additional_info_parts = [f"Persistency: {persistency:.2f}% with an overlapping weather window of {window_time} {window_input_label} and a limit of {investigation1} {('<=' if values_op == '-RADIO BELOW-' else '>=')} {limit1}"]
        else:
            persistency, count = calculate_persistency_non_overlapping(values1, window_size, limit1, values_op)
            #Mean number of occurences per year based on total number of counts by number of full months times 12
            mean_count = (count/num_full_months_approx)*12
            additional_info_parts = [f"Persistency: {persistency:.2f}% with a non-overlapping weather window of {window_time} {window_input_label} and a limit of {investigation1} {('<=' if values_op == '-RADIO BELOW-' else '>=')} {limit1}"]

        if overlapping_WW == False:
            # Append the additional provided content
            additional_info_parts.append("\n" + "\n" + f"Number of weather window occurences: {count}" + "\n"+ f"Mean number of weather window occurrences per full year: {mean_count:.2f}")

        # Join the parts to create the full additional info string
        additional_info = "; ".join(additional_info_parts)

        output_messages.append(additional_info)
        output_messages.append("")
    
    # Add a trace of the chosen variable data to the figure
    fig.add_trace(go.Scatter(x=df.index, y=df[investigation1], mode='lines', line={'color':'#1776d4', 'width': 1}, hovertemplate='%{x|%Y-%b}: %{y}<extra></extra>')) #2d9fe0

    textstr = additional_info #For the figure
    fig.add_annotation(
        x=0,
        y=0.95,
        xref='paper',
        yref='paper',
        text=textstr,
        showarrow=False,
        bgcolor='white',
        bordercolor='black',
        borderwidth=1,
        borderpad=4
    )
    # Creation of tables and output file
    start_year = df.index.min().year
    end_year = df.index.max().year
    num_years = end_year - start_year + 1

    yearly_data = []

    # For storing the sum of monthly persistencies and counts across all years
    monthly_persistencies_sum = [0] * 12
    monthly_counts_sum = [0] * 12

    for year in range(start_year, end_year + 1):
        year_data = [year]
        monthly_persistencies = []
        monthly_counts = []
        for month in range(1, 13):  # Loop through all months
            month_data = df[(df.index.month == month) & (df.index.year == year)]
            values1 = month_data[investigation1].values

            if overlapping_WW:
                persistency, count = calculate_persistency(values1, window_size, limit1, values_op)
            else:
                persistency, count = calculate_persistency_non_overlapping(values1, window_size, limit1, values_op)

            monthly_persistencies.append(persistency)
            monthly_counts.append(count)
            
            # Add persistency and count to the sum for the respective month
            monthly_persistencies_sum[month - 1] += persistency
            monthly_counts_sum[month - 1] += count

        # Calculate the yearly average persistency
        yearly_average = sum(monthly_persistencies) / len(monthly_persistencies)

        # Append monthly data and yearly average persistency to the year's data
        year_data.extend(monthly_persistencies)
        year_data.append(yearly_average)
        yearly_data.append(year_data)

    if pvalue_input == None: #This is used to indicate if an extra Pvalue was added
        pvalue_indicator = False
    else:
        pvalue_indicator = True
    # Compute the mean monthly persistencies and counts
    mean_monthly_persistencies = [persistency_sum / num_years for persistency_sum in monthly_persistencies_sum]
    monthly_data = [list(zip(*yearly_data))[i] for i in range(1, 13)]
    std_devs_per_month = [np.std(month) for month in monthly_data]
    p20_values_per_month = [np.mean(month) + std_dev * norm.ppf(1 - 0.20) for month, std_dev in zip(monthly_data, std_devs_per_month)]
    p50_values_per_month = [np.mean(month) + std_dev * norm.ppf(1 - 0.50) for month, std_dev in zip(monthly_data, std_devs_per_month)]
    p80_values_per_month = [np.mean(month) + std_dev * norm.ppf(1 - 0.80) for month, std_dev in zip(monthly_data, std_devs_per_month)]
    pvalue_input_per_month = [np.mean(month) + std_dev * norm.ppf((100 - pvalue_input)/100) for month, std_dev in zip(monthly_data, std_devs_per_month)] if pvalue_indicator == True else None
    mean_monthly_counts = [count_sum / num_years for count_sum in monthly_counts_sum]

    month_abbr = [calendar.month_abbr[i] for i in range(1, 13)]
    # Create the header and the rows for mean monthly persistencies and counts
    if overlapping_WW == False:
        header = ["Years"] + [calendar.month_abbr[i] for i in range(1, 13)] + ["Yearly Mean"]
        mean_persistencies_row = ["Monthly Mean"] + [round(persistency, 2) for persistency in mean_monthly_persistencies] + ["-"]
        mean_counts_row = ["Mean Monthly WW Occ."] + [round(count, 2) for count in mean_monthly_counts] + ["-"]
    else:
        header = ["Years"] + [calendar.month_abbr[i] for i in range(1, 13)] + ["Yearly Mean"]
        mean_persistencies_row = ["Yearly Mean"] + [round(persistency, 2) for persistency in mean_monthly_persistencies] + ["-"]

    std_dev_row = ['Std. Dev. (%)'] + [round(std_dev, 2) for std_dev in std_devs_per_month] + ['-']
    p20_row = ['P20%'] + [round(value, 2) for value in p20_values_per_month] + ['-']
    p50_row = ['P50%'] + [round(value, 2) for value in p50_values_per_month] + ['-']
    p80_row = ['P80%'] + [round(value, 2) for value in p80_values_per_month] + ['-']
    pvalue_input_row = [f'P{pvalue_input}%'] + [round(value, 2) for value in pvalue_input_per_month] + ['-'] if pvalue_indicator == True else None
    # Round data and add rounded values to formatted_data
    formatted_data = []
    for row in yearly_data:
        formatted_row = [row[0]]  # Year
        formatted_row.extend([round(val, 2) for val in row[1:-1]])  # Monthly data
        formatted_row.append(round(row[-1], 2))  # Average
        formatted_data.append(formatted_row)

    # Add the rows for mean monthly persistencies and counts
    if overlapping_WW == False:
        formatted_data.append(mean_persistencies_row)
        formatted_data.append(std_dev_row)
        formatted_data.append(p20_row)
        formatted_data.append(p50_row)
        formatted_data.append(p80_row)
        formatted_data.append(pvalue_input_row) if pvalue_indicator == True else None
        formatted_data.append(mean_counts_row)
    else:
        formatted_data.append(mean_persistencies_row)
        formatted_data.append(std_dev_row)
        formatted_data.append(p20_row)
        formatted_data.append(p50_row)
        formatted_data.append(p80_row)
        formatted_data.append(pvalue_input_row) if pvalue_indicator == True else None

    # Convert the table to a DataFrame for saving
    table_df = pd.DataFrame(formatted_data, columns=header)

    # Convert numeric columns to float type so that is represented correctly in the Excel sheet (replace None with NaN)
    numeric_cols = table_df.columns[1:-1]  # Exclude "Year" and "Yearly Average"
    table_df[numeric_cols] = table_df[numeric_cols].apply(pd.to_numeric, errors='coerce')

    # excel data saving 
    excel_output = "monthly_persistency_table.xlsx"
    table_df.to_excel(excel_output, index=False)
   
     # Creation of table to print in GUI output window
    if overlapping_WW == False: #Only adding count values for the non-overlapping case
        table_data_console = pd.DataFrame({"Month": month_abbr, "Mean Persistency": mean_monthly_persistencies, "Mean WW Occurences": mean_monthly_counts})
        table_data_console["Mean Persistency"] = table_data_console["Mean Persistency"].map("{:.2f}%".format)
        table_data_console["Mean WW Occurences"] = table_data_console['Mean WW Occurences'].map("{:.2f}".format)
        # Left-align the "Month" column and right-align the "Mean Persistency" column
        table_data_console["Month"] = table_data_console["Month"].apply(lambda x: x.ljust(14))
        table_data_console["Mean Persistency"] = table_data_console["Mean Persistency"].apply(lambda x: x.ljust(30))
        # Construct the custom header
        header = "Month".ljust(13) + "Mean Persistency".ljust(24) + "Mean WW Occurences"
    else:
        # Creation of table to print in GUI output window
        table_data_console = pd.DataFrame({"Month": month_abbr, "Mean Persistency": mean_monthly_persistencies})
        table_data_console["Mean Persistency"] = table_data_console["Mean Persistency"].map("{:.2f}%".format)
        # Left-align the "Month" column and right-align the "Mean Persistency" column
        table_data_console["Month"] = table_data_console["Month"].apply(lambda x: x.ljust(14))
        # Construct the custom header
        header = "Month".ljust(13) + "Mean Persistency"#.ljust(20)

    # Construct the table body
    table_body = table_data_console.to_string(index=False, header=False)

    # Combine the header and table body
    table_string = header + "\n" + table_body
    
    output_messages.append(table_string)

    output_messages.append("")
    output_messages.append(f"Complete persistency table has been exported to '{excel_output}' file.")

    # Add a horizontal line to the figure, the is the inputted limit
    fig.add_shape(
        type="line",
        xref="x",
        yref="y",
        x0=df.index.min(),
        y0=limit1,
        x1=df.index.max(),
        y1=limit1,
        line=dict(color="red", width=1, dash="dash")
    )

    fig.update_layout(
        title={'text': f'{investigation1} over Time', 'x':0.5, 'xanchor': 'center'},
        xaxis_title='Date',
        yaxis_title=investigation1,
        autosize=False,
        width=1400,
        height=800,
        template='none'
    )

    fig.update_xaxes(
        tickangle=45
    )

    fig.write_html('figure_persistency.html', auto_open=True) #saves and opens the figure

    return output_messages, excel_output, additional_info, overlapping_WW, pvalue_indicator
    #End of main_calculation function
   
## Running of code truly begins here

layout = [
    [sg.Text("Enter the file path for the CSV file (DHI format): "), sg.InputText(), sg.FileBrowse()],
    [sg.Button("Next"), sg.Button("Exit")]
]

window = sg.Window("Persistency Tool", layout) #First GUI window

while True:
    event, values = window.read()

    if event == "Next":
        file_path = values[0]
        window.close()
        break

    if event in (sg.WINDOW_CLOSED, "Exit"):
        break

window.close()

output_messages, excel_output, additional_info, overlapping_WW, pvalue_indicator = main_calculations(file_path, "", "") #running of the main calculation function

## BELOW HERE IS FOR THE EXCEL FILE FORMATTING

# Load the existing Excel file
wb = openpyxl.load_workbook(excel_output)

# Get the active worksheet (assuming the first sheet is active)
ws = wb.active

# Create a Border object to define grid lines
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply the Border to all cells in the worksheet
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

#AUTOSIZING
# Autosize the first and last columns, middle ones don't need it
first_column = ws['A']
last_column = ws[openpyxl.utils.get_column_letter(ws.max_column)]
columns_to_autosize = [first_column, last_column]

for column in columns_to_autosize:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column[0:]:  # Start from the first row
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) #Add two extra characters
    ws.column_dimensions[column_letter].width = adjusted_width

# Format cell values as percentages (excluding "Mean Monthly Count" row)
for row in ws.iter_rows(min_row= 2, max_row=ws.max_row - (1 if overlapping_WW == False else 0), min_col=2, max_col=ws.max_column):
    for cell in row:
            cell.number_format = '0.00"%"'

#SHADING
# Create a gray fill (shading) pattern
gray_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
# Define the light gray fill
light_gray_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

if overlapping_WW == False:
    # Apply the gray fill to persistency table
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row - (6 if pvalue_indicator == True else 5), min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = gray_fill

# Apply the light gray fill to the std dev section
    for row in ws.iter_rows(min_row=ws.max_row - (5 if pvalue_indicator == True else 4), max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = light_gray_fill

    # Apply the gray fill to the last row
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = gray_fill
else:
    # Apply the gray fill to persistency table
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row - (5 if pvalue_indicator == True else 4), min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = gray_fill

# Apply the light gray fill to the std dev section
    for row in ws.iter_rows(min_row=ws.max_row - (4 if pvalue_indicator == True else 3), max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.fill = light_gray_fill

#HEATMAP
# Determine the range for applying the heatmap
if overlapping_WW == False:
    heatmap_start_row = 6
    heatmap_start_col = 2
    heatmap_end_row = 5 + ws.max_row - (7 if pvalue_indicator == True else 6)
else:
    heatmap_start_row = 2
    heatmap_start_col = 2
    heatmap_end_row = 2 + ws.max_row - (6 if pvalue_indicator == True else 5)
heatmap_end_col = ws.max_column - 1
heatmap_range = f"{ws.cell(row=heatmap_start_row, column=heatmap_start_col).coordinate}:{ws.cell(row=heatmap_end_row, column=heatmap_end_col).coordinate}"

# Convert cell values to numeric values (assumes that cell values can be converted to float)
for row in ws.iter_rows(min_row=heatmap_start_row, max_row=heatmap_end_row, min_col=heatmap_start_col, max_col=heatmap_end_col):
    for cell in row:
        try:
            cell.value = float(cell.value)
        except (ValueError, TypeError):
            pass  # Skip non-numeric values
# Define the color scale rule for the heatmap (0% to 100%)
color_scale_rule = ColorScaleRule(start_type="num",
                                  start_value=0,
                                  start_color="FF0400",
                                  mid_type="num",
                                  mid_value=50,
                                  mid_color="FFF700",
                                  end_type="num",
                                  end_value=100,
                                  end_color="75FF00")

# Apply the heatmap to the desired range
ws.conditional_formatting.add(heatmap_range, color_scale_rule)

# Determine the range for applying the heatmap for the P values
if overlapping_WW == False:
    heatmap_pvalues_start_row = 4 + ws.max_row - (3 if pvalue_indicator == True else 2)
    heatmap_pvalues_start_col = 2
    heatmap_pvalues_end_row = 5 + ws.max_row - 1
else:
    heatmap_pvalues_start_row = 2 + ws.max_row - (3 if pvalue_indicator == True else 2)
    heatmap_pvalues_start_col = 2
    heatmap_pvalues_end_row = 2 + ws.max_row
heatmap_pvalues_end_col = ws.max_column - 1
heatmap_pvalues_range = f"{ws.cell(row=heatmap_pvalues_start_row, column=heatmap_pvalues_start_col).coordinate}:{ws.cell(row=heatmap_pvalues_end_row, column=heatmap_pvalues_end_col).coordinate}"

# Convert cell values to numeric values (assumes that cell values can be converted to float)
for row in ws.iter_rows(min_row=heatmap_pvalues_start_row, max_row=heatmap_pvalues_end_row, min_col=heatmap_pvalues_start_col, max_col=heatmap_pvalues_end_col):
    for cell in row:
        try:
            cell.value = float(cell.value)
        except (ValueError, TypeError):
            pass  # Skip non-numeric values
# Define the color scale rule for the heatmap (0% to 100%)
color_scale_rule = ColorScaleRule(start_type="num",
                                  start_value=0,
                                  start_color="FF0400",
                                  mid_type="num",
                                  mid_value=50,
                                  mid_color="FFF700",
                                  end_type="num",
                                  end_value=100,
                                  end_color="75FF00")

# Apply the heatmap to the desired range
ws.conditional_formatting.add(heatmap_pvalues_range, color_scale_rule)

# Insert additional_info at the beginning of the worksheet
if overlapping_WW == False:
    # Insert additional_info at the beginning of the worksheet
    ws.insert_rows(1)
    ws.insert_rows(2)
    ws.insert_rows(3)
    ws.insert_rows(4)
    additional_info_lines = additional_info.split('\n')
    for i, line in enumerate(additional_info_lines):
        ws.cell(row=i + 1, column=1, value=line)
    ws.insert_rows(5)
else:
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=additional_info)
    ws.insert_rows(2)
# Save the modified Excel file with additional_info and formatted table
wb.save(excel_output)


# Determine the number of lines and the maximum line length in the output
lines = "\n".join(output_messages).split("\n")
num_lines = len(lines)
max_line_length = max(len(line) for line in lines)

# Create a multiline widget with a dynamic size
output_window = sg.Window('Output', [[sg.Multiline("\n".join(output_messages), size=(max_line_length, num_lines))], [sg.Button('Exit Program')]])

#Create and open final GUI window
while True:
    event, values = output_window.read()
    if event == sg.WINDOW_CLOSED or event == 'Exit Program':
        break
output_window.close() #Program ends when final GUI window is closed