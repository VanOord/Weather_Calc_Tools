import pandas as pd
import numpy as np
import os
import calendar
import PySimpleGUI as sg
from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import webbrowser
import time

def generate_scatter_table_v2(start_month_num=None, end_month_num=None):
    # Filter data based on the month range if provided
    filtered_data = df
    if start_month and end_month:
        filtered_data = df[(df['Month'] >= start_month_num) & (df['Month'] <= end_month_num)]
    
    # Create a pivot table to calculate the count
    pivot_table_filtered = pd.pivot_table(filtered_data, values=datetime, index='Scatter Table', columns='Tp Bin', aggfunc='count', fill_value=0)
    
    # Calculate the total count
    total_count_filtered = pivot_table_filtered.sum().sum()

    # Calculate the percentage of occurrence
    percentage_table_filtered = (pivot_table_filtered / total_count_filtered) * 100
    
    # Calculate total and cumulative percentages for rows
    percentage_table_filtered['Total Row'] = percentage_table_filtered.sum(axis=1)
    percentage_table_filtered['Accum Row'] = percentage_table_filtered['Total Row'].cumsum()
    
    # Calculate total and cumulative percentages for columns
    total_col = percentage_table_filtered.sum(axis=0)
    accum_col = total_col.cumsum()
    percentage_table_filtered.loc['Total Col', :] = total_col
    percentage_table_filtered.loc['Accum Col', :] = accum_col

    # Round all values to 3 decimal places
    percentage_table_filtered = percentage_table_filtered.round(3)

    # Set the intersection of "Total" and "Accum" rows and columns to "-"
    percentage_table_filtered.loc['Total Col', 'Total Row'] = '-'
    percentage_table_filtered.loc['Total Col', 'Accum Row'] = '-'
    percentage_table_filtered.loc['Accum Col', 'Total Row'] = '-'
    percentage_table_filtered.loc['Accum Col', 'Accum Row'] = '-'

    return percentage_table_filtered

def apply_heatmap_to_data(ws):
    """
    Apply a heatmap to the main data range in the worksheet, excluding "Accum" and "Total" rows and columns.
    """
    # Define the range for applying the heatmap
    # Starting from C5, but the end cell will exclude "Accum" and "Total" rows and columns
    max_row = ws.max_row - 2  # Subtracting 2 to exclude "Accum" and "Total" rows
    max_col = ws.max_column - 2  # Subtracting 2 to exclude "Accum" and "Total" columns
    cell_range = f"C5:{ws.cell(row=max_row, column=max_col).coordinate}"

    # Define the color scale rule for the heatmap
    color_scale_rule = ColorScaleRule(start_type="min",
                                      start_color="FFFFFF",
                                      mid_type="percentile",
                                      mid_value=25,
                                      mid_color="C3E7FF",
                                      end_type="max",
                                      end_color="5D849F")

    # Apply the heatmap to the desired range
    ws.conditional_formatting.add(cell_range, color_scale_rule)

# Define a function to add gridlines to a worksheet
def apply_gridlines(ws):
    # Define the desired gridline style: gray, thinly dotted
    thin_dotted_gray = Side(border_style="dashed", color="000000") #dotted  B5B5B5

    # Define the desired border style: standard black
    standard_black = Side(border_style="thin", color="000000")

    # Define the range for applying the gridlines (only to main_data)
    max_row_main_data = ws.max_row - 2  # Subtracting 2 to exclude "Accum" and "Total" rows
    max_col_main_data = ws.max_column - 2  # Subtracting 2 to exclude "Accum" and "Total" columns

    # Apply gridlines to the main_data cells
    for row in ws.iter_rows(min_row=5, max_row=max_row_main_data, min_col=3, max_col=max_col_main_data):
        for cell in row:
            cell.border = Border(
                top=thin_dotted_gray if cell.border.top.style is None else cell.border.top,
                left=thin_dotted_gray if cell.border.left.style is None else cell.border.left,
                right=thin_dotted_gray if cell.border.right.style is None else cell.border.right,
                bottom=thin_dotted_gray if cell.border.bottom.style is None else cell.border.bottom
            )
    
    border = Border(top=standard_black, left=standard_black, right=standard_black, bottom=standard_black)

    # Define the range for applying the gridlines to "Total" and "Accum" rows and columns
    max_row = ws.max_row
    max_col = ws.max_column

    # Apply border to the last two rows ("Total" and "Accum" rows)
    for row in ws.iter_rows(min_row=max_row-1, max_row=max_row, min_col=3, max_col=max_col):
        for cell in row:
            cell.border = border

    # Apply border to the last two columns ("Total" and "Accum" columns)
    for col in ws.iter_cols(min_col=max_col-1, max_col=max_col, min_row=5, max_row=max_row):
        for cell in col:
            cell.border = border

def apply_gray_header(ws):
    # Define the desired fill: gray
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

     # Apply gray fill to the header column starting from column 2 (excluding B2, B3, and B4)
    for cell in ws['B']:
        if cell.row not in [1, 2, 3]:
            cell.fill = gray_fill
    
    # Apply gray fill to the header row starting from row 4 (excluding cell A4)
    for cell in ws[4]:
        if cell.column != 1:  # Exclude column A (cell A4)
            cell.fill = gray_fill
    
    pink_fill = PatternFill(start_color="FFDAAF", end_color="FFDAAF", fill_type="solid")
     # Loop through the cells where scatter_table_v2 is written and apply blue shading to bins below and including (4.75, 5.0]
    start_shading = False  # Flag to determine when to start shading
    for row in range(5, ws.max_row + 1):  # Assuming data starts from row 5
        for col in range(2, ws.max_column + 1):  # Assuming data starts from column 2
            cell = ws.cell(row=row, column=col)
            if cell.value == "(4.75, 5.0]":
                start_shading = True  # Start shading from this cell onward
            if start_shading:
                cell.fill = pink_fill

def autosize_cells(ws):
    # Autosize columns based on the content of respective cells in row 4
    for idx in range(1, ws.max_column):
        cell_value = ws.cell(row=4, column=idx).value
        if cell_value:
            adjusted_width = len(str(cell_value))  # Add a little extra width for clarity
            ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    ws.column_dimensions['A'].width = 3 #column_1_width

    # Adjust the width of the last two columns, "Total Row" and "Accum Row", based on their content
    for idx in range(ws.max_column - 1, ws.max_column + 1):  # Only the last two columns
        max_length = max([len(str(ws.cell(row=row_num, column=idx).value)) for row_num in range(1, ws.max_row + 1) if ws.cell(row=row_num, column=idx).value])
        adjusted_width = max_length + 2  # Add a little extra width for clarity
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

def add_title_and_axis_labels(ws, start_month, end_month, month_label, title="Title", x_label="X Axis", y_label="Y Axis"):
    # Insert three rows at the beginning and one column at the start
    ws.insert_rows(idx=1, amount=3)
    ws.insert_cols(idx=1)

    # Set the title in the first row, spanning across all columns
    title_cell = ws.cell(row=1, column=2, value=title)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=ws.max_column)
    title_cell.font = Font(bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the y-axis label in the fourth row, centered across all columns
    y_label_cell = ws.cell(row=3, column=2, value=y_label)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=ws.max_column)
    y_label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the x-axis label in the first column, centered vertically across all rows
    x_label_cell = ws.cell(row=3, column=1, value=x_label)
    ws.merge_cells(start_row=3, start_column=1, end_row=ws.max_row, end_column=1)
    x_label_cell.alignment = Alignment(horizontal="center", vertical="center", text_rotation = 90)

    #Add covered months label in cell B2
    if investigate_range:
        month_label = f"Covered months: {start_month} through {end_month}"
    else:    
        month_label = "Covered months: All"
    month_label_cell = ws.cell(row=2, column=2, value=month_label)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
    month_label_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Define the desired fill: gray
    gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEDAD", end_color="FFEDAD", fill_type="solid")
    
    # Apply gray shading to the title, x-label, and y-label
    title_cell.fill = gray_fill
    y_label_cell.fill = gray_fill
    x_label_cell.fill = gray_fill
    month_label_cell.fill = yellow_fill

    # Set the CSV filename in cell F2
    ws['F2'].value = f"Data taken from: {filename}"

def shade_total_accum_cells(ws):
    """
    Shade the "Total" and "Accum" columns and rows with a light gray background.
    """
    # Define the light gray fill
    light_gray_fill = PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")
    
    # Get the last two columns (assuming they are "Total" and "Accum")
    last_col = ws.max_column
    end_col_shade = last_col - 1
    
    # Get the last two rows (assuming they are "Total" and "Accum")
    last_row = ws.max_row
    end_row_shade = last_row - 1
    
    # Shade the last two columns
    for row in ws.iter_rows(min_col=end_col_shade, max_col=last_col, min_row=4):
        for cell in row:
            cell.fill = light_gray_fill

    # Shade the last two rows
    for col in ws.iter_cols(min_row=end_row_shade, max_row=last_row):
        for cell in col:
            cell.fill = light_gray_fill

def find_data_start(file_path):
    with open(file_path) as file:
        lines = file.readlines()

    for i, line in enumerate(lines):
        row_data = line.strip().split(",")
        if len(row_data) >= 2:
            try:
                parse(row_data[0])
                if all(len(line.strip().split(",")) >= 2 for line in lines[i+1:i+21]):
                    return i - 1  # Adjusted to return the index of the header line
            except ValueError:
                continue

    return None
##CODE STARTS HERE
layout = [
    [sg.Text("Enter the file path for the CSV file (DHI format): "), sg.InputText(), sg.FileBrowse()],
    [sg.Button("Next"), sg.Button("Exit")]
]

window = sg.Window("Scatter Table Tool", layout)

while True:
    event, values = window.read()

    if event == "Next":
        file_path = values[0]
        # Extract filename from the full path
        filename = os.path.basename(file_path)
        window.close()
        break

    if event in (sg.WINDOW_CLOSED, "Exit"):
        break

window.close()

# Determine the start of the data based on if there is a datedatetime format atleast 20 rows long in the first column
data_start = find_data_start(file_path)

# Load chunk of data from the CSV file to get the header names of the columns
chunk_df = pd.read_csv(file_path, nrows=100, skiprows=data_start)
datetime = chunk_df.columns[0] #will be the index column, CODE ONLY WORKS IF DATEdatetime COLUMN IS FIRST
header_names1 = list(chunk_df.columns[1:]) #Skips first column header (datetime)
header_names2 = list(chunk_df.columns[2:]) #Skips first two column headers


# Define the layout for the GUI window
layout = [
        [sg.Text("Choose columns to investigate: ")],
        [sg.Text("Column 1: "), sg.Combo(header_names1, key='investigation1'), sg.Text("Limit: ")],
        [sg.Text("Column 2: "), sg.Combo(header_names2, key='investigation2'), sg.Text("Limit: ")],
        [sg.Text('Do you want to create scatter table based on certain months?')],
        [sg.Radio('No', 'RADIO1', key='-RADIO NO-', default=True), sg.Radio('Yes', 'RADIO1', key='-RADIO YES-')],
        [sg.Text('              Start month:'), sg.Combo(list(calendar.month_name[1:]), key='-START MONTH-', enable_events=True, disabled=False)],
        [sg.Text('              End month:'), sg.Combo(list(calendar.month_name[1:]), key='-END MONTH-', enable_events=True, disabled=False)],
        [sg.Button("Plot"), sg.Button("Exit")]
    ]

# Create the GUI window
window = sg.Window("Scatter Table Tool", layout)

# Event loop to process GUI events
while True:
    event, values = window.read()

    if event == "Plot":
            investigation1 = values['investigation1']
            investigation2 = values['investigation2']
            investigate_range = values['-RADIO YES-']
            start_month = values['-START MONTH-']
            end_month = values['-END MONTH-']
            window.close()
            break

    if event in (sg.WINDOW_CLOSED, "Exit"):
        break

# Close the GUI window
window.close()

# Create a list of columns to load based on user's input
columns_to_load = [datetime, investigation1, investigation2]

# Load the CSV with only the selected columns
df = pd.read_csv(file_path, skiprows=data_start, usecols=columns_to_load, parse_dates=[datetime], dayfirst=True)

df_datetime = df[datetime]
# Extract month from the datetime column
df['Month'] = df_datetime.dt.month

x_bin = 1
y_bin = 0.5
height_bins_025 = np.arange(0, 5, 0.25)
height_bins_05 = np.arange(5, df[investigation1].max() + y_bin, y_bin)
height_bins = np.concatenate([height_bins_025, height_bins_05])
period_bins = np.arange(0, df[investigation2].max() + x_bin, x_bin)

# Bin the wave height and wave period values
df['Scatter Table'] = pd.cut(df[investigation1], bins=height_bins)
df['Tp Bin'] = pd.cut(df[investigation2], bins=period_bins)

if investigate_range: #this case is for if certain months want to be investigated
    month_dict = {v.lower(): k for k, v in enumerate(calendar.month_name) if k != 0}
    start_month_num = month_dict[start_month.lower()]
    end_month_num = month_dict[end_month.lower()]
else:
    start_month_num = None
    end_month_num = None

# Generate scatter table
scatter_table_v2 = generate_scatter_table_v2(start_month_num, end_month_num)

# Split the DataFrame into two parts
main_data = scatter_table_v2.iloc[:-2]  # Excluding the last two rows (Accum and Total)
accum_total_rows = scatter_table_v2.iloc[-2:]

# Reverse the order of rows in the main data
reversed_main_data = main_data.iloc[::-1]

# Concatenate the reversed main data with the Accum and Total rows
scatter_table_v2 = pd.concat([reversed_main_data, accum_total_rows])
try:
    saved_file_name = f'Scatter Table {investigation1} vs {investigation2}.xlsx'
    scatter_table_v2.to_excel(saved_file_name)
except Exception as e:
    saved_file_name = "Scatter Table Output.xlsx"
    scatter_table_v2.to_excel(saved_file_name)

# Load the workbook and select the sheet
wb = load_workbook(saved_file_name)
ws = wb.active

add_title_and_axis_labels(ws, start_month, end_month, "", title=(f"Frequency of Occurence [%] for {investigation1} vs {investigation2}"), x_label=f"{investigation1}", y_label=f"{investigation2}")
apply_heatmap_to_data(ws)
apply_gridlines(ws)
apply_gray_header(ws)
autosize_cells(ws)
shade_total_accum_cells(ws)
# Save the modified workbook
wb.save(saved_file_name)
webbrowser.open(saved_file_name)
# Display the completion popup message
#sg.Popup(f"Processing complete. Data saved to Scatter table {investigation1} vs {investigation2}.xlsx")
# Introduce a delay
time.sleep(5)  # waits for 10 seconds ##THIS LINE WILL NOT WORK WITH VSCODE, ONLY IN CMD